import io
import re
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import requests
import streamlit as st
from fpdf import FPDF
from openpyxl import load_workbook

APP_VERSION = "V65 Ijazah Checklist Guard"
DEFAULT_REFRESH_SECONDS = 60

st.set_page_config(page_title="HR Data Filter", page_icon="🔎", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif !important;
    color: #0F172A !important;
}
[data-testid="stAppViewContainer"] {
    background: linear-gradient(180deg, #F8FAFC 0%, #EEF4FA 100%);
}
[data-testid="stHeader"] {
    background: #F8FAFC !important;
}
.block-container {
    padding-top: 2.2rem !important;
    padding-bottom: 2.4rem !important;
    max-width: 1380px !important;
}

/* Header */
.main-header {
    background: #FFFFFF;
    border: 1px solid #D6E0EA;
    border-left: 7px solid #0F4C81;
    border-radius: 20px;
    padding: 30px 34px;
    margin-bottom: 22px;
    box-shadow: 0 14px 34px rgba(15, 23, 42, .07);
}
.header-title {
    font-size: 33px;
    font-weight: 800;
    color: #0B1B33;
    margin: 0;
    letter-spacing: -.035em;
}
.header-subtitle {
    font-size: 15px;
    color: #475569;
    margin-top: 8px;
    font-weight: 600;
}

/* Cards / sections */
.section-card {
    background: #FFFFFF;
    border: 1px solid #D6E0EA;
    border-radius: 18px;
    padding: 24px;
    margin-bottom: 20px;
    box-shadow: 0 10px 30px rgba(15, 23, 42, .055);
    overflow: visible !important;
}
.section-title {
    font-size: 22px;
    font-weight: 800;
    color: #0B1B33;
    margin-bottom: 6px;
}
.section-caption {
    font-size: 13px;
    color: #64748B;
    margin-bottom: 16px;
    font-weight: 700;
}

/* KPI only 2 cards now */
.metric-grid {
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 16px;
    margin-bottom: 20px;
}
.metric-card {
    background: #FFFFFF;
    border: 1px solid #D6E0EA;
    border-radius: 16px;
    padding: 20px 22px;
    box-shadow: 0 10px 28px rgba(15,23,42,.055);
    min-height: 122px;
}
.metric-label {
    font-size: 12px;
    color: #31516F;
    text-transform: uppercase;
    letter-spacing: .09em;
    font-weight: 800;
}
.metric-value {
    margin-top: 10px;
    font-size: 29px;
    font-weight: 800;
    color: #0F4C81;
    letter-spacing: -.03em;
}
.metric-note {
    margin-top: 5px;
    font-size: 13px;
    color: #64748B;
    font-weight: 600;
}

/* Inputs */
div[data-baseweb="input"], div[data-baseweb="select"] {
    overflow: visible !important;
}
.stTextInput, .stSelectbox, .stMultiSelect {
    margin-bottom: 12px !important;
    overflow: visible !important;
}
.stTextInput input,
.stSelectbox div[data-baseweb="select"] > div,
.stMultiSelect div[data-baseweb="select"] > div {
    min-height: 52px !important;
    border-radius: 12px !important;
    border: 1px solid #B8C8DA !important;
    box-shadow: none !important;
    background: #FFFFFF !important;
    color: #0F172A !important;
    font-size: 15px !important;
}
.stTextInput input:focus,
.stSelectbox div[data-baseweb="select"] > div:focus-within {
    border-color: #0F4C81 !important;
    box-shadow: 0 0 0 3px rgba(15, 76, 129, .12) !important;
}
label, .stTextInput label, .stSelectbox label {
    color: #0F172A !important;
    font-weight: 750 !important;
}

/* Buttons */
.stButton > button, .stDownloadButton > button {
    min-height: 52px !important;
    border-radius: 12px !important;
    font-weight: 800 !important;
    border: 1px solid #B8C8DA !important;
    background: #FFFFFF !important;
    color: #0F172A !important;
    font-size: 15px !important;
}
.stButton > button:hover, .stDownloadButton > button:hover {
    border-color: #0F4C81 !important;
    color: #0F4C81 !important;
}
button[kind="primary"] {
    background: #0F4C81 !important;
    color: #FFFFFF !important;
    border-color: #0F4C81 !important;
}

/* Dataframe */
div[data-testid="stDataFrame"] {
    border-radius: 16px !important;
    overflow: hidden !important;
    border: 1px solid #D6E0EA !important;
    background: #FFFFFF !important;
}
.notice-box {
    background:#FFF7ED;
    border:1px solid #FED7AA;
    color:#9A3412;
    padding:14px 16px;
    border-radius:14px;
    font-weight:700;
    margin-bottom:14px;
}
hr { display: none !important; }

@media(max-width:900px){
    .metric-grid{grid-template-columns:1fr;}
    .header-title{font-size:27px;}
}
</style>
""", unsafe_allow_html=True)

BLACKLIST = ["kode", "data pendukung", "catatan"]
EMPTY = {"", "nan", "none", "null", "-", "--", "belum ada", "belum ada di db", "tidak ada"}
CHECK = {"v", "√", "✓", "check", "checked", "ya", "yes", "ada", "true", "1"}
MONTHS = {"januari":1,"jan":1,"februari":2,"feb":2,"maret":3,"mar":3,"april":4,"apr":4,"mei":5,"juni":6,"jun":6,"juli":7,"jul":7,"agustus":8,"agust":8,"agu":8,"aug":8,"september":9,"sept":9,"sep":9,"oktober":10,"okto":10,"okt":10,"november":11,"nov":11,"desember":12,"des":12}
MONTH_ID = {1:"Januari",2:"Februari",3:"Maret",4:"April",5:"Mei",6:"Juni",7:"Juli",8:"Agustus",9:"September",10:"Oktober",11:"November",12:"Desember"}
MAIN_COLS = ["NO","NAMA","LINK PERSONIL","STRATA","JENIS IJAZAH","TAHUN LULUS IJAZAH","KEAHLIAN","PENERBIT SKA/SKK","BERLAKU SKA","TGL EXPIRED SKA","STATUS SKA","STATUS KERJA","KOTA/KABUPATEN","PROVINSI","NO NIK","NO NPWP","NO. TELP","EMAIL","SUMBER","KATEGORI_ASAL"]
DEFAULT_COLS = ["NO","NAMA","LINK PERSONIL","JENIS IJAZAH","KEAHLIAN","TGL EXPIRED SKA","STATUS SKA","STATUS KERJA","KOTA/KABUPATEN","PROVINSI","TAHUN LULUS IJAZAH"]


def clean(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)): return ""
    s = re.sub(r"\s+", " ", str(x).replace("\xa0", " ").replace("\n", " ")).strip()
    if s.lower() in EMPTY: return ""
    if re.fullmatch(r"\d+\.0", s): s = s[:-2]
    return s


def norm_header(x) -> str:
    s = clean(x).upper().replace("/", " / ").replace(".", " ").replace("(", " ").replace(")", " ")
    return re.sub(r"\s+", " ", s).strip()


def unique_cols(cols):
    seen, out = {}, []
    for c in cols:
        b = norm_header(c) or "UNNAMED"
        seen[b] = seen.get(b, 0) + 1
        out.append(b if seen[b] == 1 else f"{b}__{seen[b]}")
    return out


def classify(col: str) -> Optional[str]:
    base = re.sub(r"__\d+$", "", norm_header(col))
    c = re.sub(r"[^A-Z0-9]", "", base)
    if c in {"NO","NOMOR","NOURUT","INDEX"}: return "NO"
    if c == "NAMA": return "NAMA"
    if c == "STRATA": return "STRATA"
    if "KEAHLIAN" in c or "SKASKKAKTIF" in c or "SKKYANGDIMILIKI" in c: return "KEAHLIAN"
    if "JENISIJAZAH" in c or "JENISIJASAH" in c or c in {"IJAZAH","IJASAH","PENDIDIKAN"} or "IJASADANKELULUSAN" in c: return "JENIS IJAZAH"
    if "TAHUNLULUS" in c or "TAHUNSERTIFIKAT" in c: return "TAHUN LULUS IJAZAH"
    if c in {"STATUSKERJA", "STATUSPERSONIL", "STATUSKARYAWAN"} or base == "STATUS KERJA": return "STATUS KERJA"
    if "PENERBITSKA" in c or c == "PENERBIT": return "PENERBIT SKA/SKK"
    if "BERLAKUSKA" in c: return "BERLAKU SKA"
    if "TGLEXPIREDSKA" in c or "EXPIREDSKA" in c: return "TGL EXPIRED SKA"
    if "KOTA" in c and ("PROVINSI" in c or "PROPINSI" in c): return "DOMISILI"
    if "DOMISILI" in c: return "DOMISILI"
    if c in {"PROPINSI","PROVINSI"} or c.startswith("PROPINSI") or c.startswith("PROVINSI"): return "PROVINSI"
    if "KABUPATEN" in c or c in {"KOTA","KABKOTA","KOTAKAB"} or "KOTAKAB" in c: return "KOTA/KABUPATEN"
    if c in {"SUMBER","SOURCE"}: return "SUMBER"
    if c in {"EMAIL","EEMAIL"}: return "EMAIL"
    if "TELP" in c or "TELEPON" in c or c in {"HP","HANDPHONE"}: return "NO. TELP"
    if "NONIK" in c or c == "NIK": return "NO NIK"
    if "NONPWP" in c: return "NO NPWP"
    if c == "NPWP": return "NPWP_RAW"
    if "PENGALAMAN" in c: return "PENGALAMAN KERJA (TAHUN)"
    if c in {"KETERANGAN","CATATAN"}: return "KETERANGAN"
    return None


def first(row, cols):
    for col in cols:
        v = clean(row.get(col, ""))
        if v: return v
    return ""


def first_education(row, cols):
    """Return first meaningful education value and skip document checklist marks.

    In the source workbook, standalone IJASAH/IJAZAH can mean document checklist,
    while JENIS IJAZAH / IJASA DAN KELULUSAN means education. Because headers are
    inconsistent, both may be grouped as JENIS IJAZAH. This helper ignores values
    like v/√ so they do not appear in the education column.
    """
    for col in cols:
        v = clean(row.get(col, ""))
        if not v or is_check(v):
            continue
        if email_of(v) or phone_of(v) or nik_of(v) or npwp_of(v):
            continue
        return v
    return ""


def digits(s): return re.sub(r"\D", "", clean(s))
def is_check(s): return clean(s).lower() in CHECK

def is_dateish(s):
    t = clean(s).lower()
    if not t: return False
    if re.search(r"\b\d{4}-\d{1,2}-\d{1,2}(\s+00:00:00)?\b", t): return True
    if any(m in t for m in MONTHS) and re.search(r"\d", t): return True
    if re.fullmatch(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", t): return True
    return False


def email_of(s):
    m = re.search(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", clean(s), re.I)
    return m.group(0) if m else ""


def source_clean(s):
    s = clean(s)
    if not s or is_check(s): return ""
    if "mailto:" in s.lower() and email_of(s): return ""
    m = re.search(r'\(["\'].*?["\']\s*,\s*["\']([^"\']+)["\']\)', s)
    if m:
        display = clean(m.group(1))
        return display if display and not email_of(display) and not display.lower().startswith("http") else ""
    if s.lower().startswith("http"): return ""
    return s


def phone_of(s):
    s = clean(s)
    if not s or is_dateish(s) or re.search(r"[A-Za-z]", s): return ""
    d = digits(s)
    if not 9 <= len(d) <= 14: return ""
    if d.startswith("08"): return s
    if d.startswith("628"): return "+" + d if not s.startswith("+") else s
    if d.startswith("8") and len(d) >= 9: return "0" + d
    return ""


def nik_of(s):
    d = digits(s)
    return d if len(d) == 16 else ""


def npwp_of(s):
    s = clean(s); d = digits(s)
    return s if len(d) in {15, 16} and ("." in s or "-" in s or "/" in s) else ""


def norm_filter(s):
    s = clean(s).lower().replace("/", " ").replace("-", " ").replace(".", " ").replace(",", " ")
    return re.sub(r"\s+", " ", s).strip()


def norm_education(s):
    """Normalize education text so D IV/D-IV/D.IV/Diploma IV can match D4."""
    s = norm_filter(s)
    replacements = [
        (r"\bd\s*iv\b", "d4"),
        (r"\bdiploma\s*iv\b", "d4"),
        (r"\bd\s*iii\b", "d3"),
        (r"\bdiploma\s*iii\b", "d3"),
        (r"\bd\s*ii\b", "d2"),
        (r"\bdiploma\s*ii\b", "d2"),
        (r"\bd\s*i\b", "d1"),
        (r"\bdiploma\s*i\b", "d1"),
        (r"\bs\s*1\b", "s1"),
        (r"\bs\s*2\b", "s2"),
        (r"\bs\s*3\b", "s3"),
    ]
    for pat, repl in replacements:
        s = re.sub(pat, repl, s, flags=re.I)
    return re.sub(r"\s+", " ", s).strip()


def split_terms(s):
    s = clean(s)
    if not s: return []
    parts = re.split(r"\s*(?:,|;|\||/|\batau\b|\bdan\b|\bserta\b|\bor\b)\s*", s, flags=re.I)
    terms = [norm_filter(p) for p in parts if norm_filter(p)]
    edu = re.findall(r"\b(?:SMA|SMK|D1|D2|D3|D4|S1|S2|S3)\b", s.upper())
    if len(edu) >= 2 and len(terms) <= 1: terms = [x.lower() for x in edu]
    return list(dict.fromkeys(terms))


def simplify_skill(s):
    s = norm_filter(s)
    for w in ["skk","ska","sertifikat","ahli","teknik","jenjang","utama","madya","muda"]:
        s = re.sub(rf"\b{w}\b", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def contains_any(series, terms):
    if not terms: return pd.Series([True] * len(series), index=series.index)
    h = series.fillna("").astype(str).map(norm_filter)
    mask = pd.Series([False] * len(series), index=series.index)
    for term in terms:
        if term:
            mask |= h.str.contains(re.escape(term), na=False)
    return mask


def split_query_groups(s):
    """OR groups separated by comma/serta/dan/atau, AND tokens inside each group.

    Example:
    - "S1 serta D3" -> [["s1"], ["d3"]]
    - "Sipil D4" -> [["sipil", "d4"]]
    - "Manado, Sulut, Bali" -> [["manado"], ["sulut"], ["bali"]]
    """
    s = clean(s)
    if not s:
        return []
    groups = re.split(r"\s*(?:,|;|\||/|\batau\b|\bdan\b|\bserta\b|\bor\b)\s*", s, flags=re.I)
    out = []
    for group in groups:
        norm = norm_filter(group)
        tokens = [t for t in norm.split() if t]
        if tokens:
            out.append(tokens)
    return out


def contains_groups(series, query):
    groups = split_query_groups(query)
    if not groups:
        return pd.Series([True] * len(series), index=series.index)
    h = series.fillna("").astype(str).map(norm_filter)
    final = pd.Series([False] * len(series), index=series.index)
    for tokens in groups:
        group_mask = pd.Series([True] * len(series), index=series.index)
        for token in tokens:
            group_mask &= h.str.contains(re.escape(token), na=False)
        final |= group_mask
    return final



EDU_LEVEL_RE = re.compile(r"\b(SMA|SMK|D1|D2|D3|D4|S1|S2|S3)\b", re.I)

def extract_edu_levels(text_value):
    return list(dict.fromkeys(x.lower() for x in EDU_LEVEL_RE.findall(norm_education(text_value))))


def build_education_helpers(jenis_ijazah, strata):
    """Build education search fields with strict source priority.

    Rule:
    1. JENIS IJAZAH is the primary source for education filtering.
    2. STRATA is used only when JENIS IJAZAH is empty.
    3. Checklist marks in JENIS IJAZAH are treated as empty.
    """
    jenis_clean = clean(jenis_ijazah)
    strata_clean = clean(strata)

    if is_check(jenis_clean):
        jenis_clean = ""
    if is_check(strata_clean):
        strata_clean = ""

    jenis_norm = norm_education(jenis_clean)
    strata_norm = norm_education(strata_clean)

    source = jenis_norm if jenis_norm else strata_norm
    levels = extract_edu_levels(source)

    return source, " ".join(levels)

def contains_education_query(df_or_series, query):
    """Smart and strict education filter.

    Supported:
    - "sipil d4"      => major sipil AND level d4
    - "sipil d4 d3"   => (major sipil AND level d4) OR (major sipil AND level d3)
    - "s1, d3"        => level s1 OR level d3
    - "s1 serta d3"   => level s1 OR level d3

    This function accepts either a dataframe with helper columns or a series.
    Dataframe mode is preferred because it can use strict _EDU_LEVELS.
    """
    q = clean(query)
    if not q:
        if isinstance(df_or_series, pd.DataFrame):
            return pd.Series([True] * len(df_or_series), index=df_or_series.index)
        return pd.Series([True] * len(df_or_series), index=df_or_series.index)

    if isinstance(df_or_series, pd.DataFrame):
        h = df_or_series.get("_PENDIDIKAN_SEARCH", pd.Series([""] * len(df_or_series), index=df_or_series.index)).astype(str)
        level_series = df_or_series.get("_EDU_LEVELS", pd.Series([""] * len(df_or_series), index=df_or_series.index)).astype(str)
    else:
        h = df_or_series.fillna("").astype(str).map(norm_education)
        level_series = h

    q_norm = norm_education(q)
    levels = [x.lower() for x in EDU_LEVEL_RE.findall(q_norm)]
    has_explicit_or = bool(re.search(r",|;|\||/|\batau\b|\bdan\b|\bserta\b|\bor\b", q, re.I))

    # Case: "sipil d4 d3" without comma.
    # Interpret as (sipil+d4) OR (sipil+d3), not sipil+d4+d3 in one row.
    if len(set(levels)) >= 2 and not has_explicit_or:
        major_text = EDU_LEVEL_RE.sub(" ", q_norm)
        major_tokens = [t for t in major_text.split() if t]
        final = pd.Series([False] * len(h), index=h.index)
        for level in dict.fromkeys(levels):
            mask = level_series.str.contains(r"(^|\s)" + re.escape(level) + r"(\s|$)", na=False, regex=True)
            for token in major_tokens:
                mask &= h.str.contains(re.escape(token), na=False)
            final |= mask
        return final

    # Normal grouped matching.
    groups = split_query_groups(q)
    if not groups:
        return pd.Series([True] * len(h), index=h.index)

    final = pd.Series([False] * len(h), index=h.index)
    for tokens in groups:
        group_levels = [t for t in tokens if EDU_LEVEL_RE.fullmatch(t.upper())]
        major_tokens = [t for t in tokens if t not in group_levels]
        group_mask = pd.Series([True] * len(h), index=h.index)

        for level in group_levels:
            group_mask &= level_series.str.contains(r"(^|\s)" + re.escape(level) + r"(\s|$)", na=False, regex=True)
        for token in major_tokens:
            group_mask &= h.str.contains(re.escape(token), na=False)

        final |= group_mask

    return final


def contains_global(df, q):
    q = norm_filter(q)
    if not q: return pd.Series([True] * len(df), index=df.index)
    h = df.get("_SEARCH_TEXT", pd.Series([""] * len(df), index=df.index)).astype(str)
    mask = pd.Series([True] * len(df), index=df.index)
    for token in q.split(): mask &= h.str.contains(re.escape(token), na=False)
    return mask


def parse_date(s):
    s = clean(s)
    if not s: return pd.NaT
    if re.fullmatch(r"\d+(\.0)?", s):
        try:
            n = float(s)
            if 20000 <= n <= 60000: return pd.to_datetime(n, unit="D", origin="1899-12-30", errors="coerce")
        except Exception: pass
    s2 = s
    if re.search(r"\bs\.?d\.?\b", s2, re.I): s2 = re.split(r"\bs\.?d\.?\b", s2, flags=re.I)[-1]
    elif " - " in s2: s2 = s2.split(" - ")[-1]
    m = re.search(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", s2.lower())
    if m:
        month = MONTHS.get(m.group(2).lower())
        if month: return pd.Timestamp(year=int(m.group(3)), month=month, day=int(m.group(1)))
    return pd.to_datetime(s2, errors="coerce", dayfirst=True)


def show_date(s):
    s = clean(s)
    if not s: return ""
    # Keep pure year values as years, not "1 Januari YYYY".
    if re.fullmatch(r"(?:19\d{2}|20\d{2})(?:\s*;\s*(?:19\d{2}|20\d{2}))*", s): return s
    if re.search(r"\b(S1|S2|S3|D3|D4)\b", s, re.I) and "," in s: return s.replace(" 00:00:00", "")
    dt = parse_date(s)
    return f"{dt.day} {MONTH_ID[int(dt.month)]} {dt.year}" if pd.notna(dt) else s.replace(" 00:00:00", "")

def extract_years_from_education_text(s):
    """Extract likely graduation years from combined education text.

    Examples:
    - "S1 Sipil 2017" -> "2017"
    - "D4 Sipil 2021 S2 Sipil 2025" -> "2021, 2025"
    """
    s = clean(s)
    years = re.findall(r"\b(?:19|20)\d{2}\b", s)
    return ", ".join(dict.fromkeys(years))


def clean_education_label(s):
    """Normalize education label and remove checklist-only values.

    Example:
    - "S1 Sipil 2017" -> "S1 Sipil"
    - "√" / "v" -> ""
    """
    s = clean(s)
    if not s or is_check(s):
        return ""
    s = re.sub(r"\b(?:19|20)\d{2}\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip(" ,;-/")
    return "" if is_check(s) else s

def latest_expiry_from_text(s):
    """Extract the latest expiry-like date from free-text SKA/SKK descriptions.

    TAT ESC/BPS often stores SKA/SKK information as text, e.g.
    "- STRA Madya Exp : 10 April 2028 - SKK Arsitek Madya Exp Oktober 2030".
    This function finds date candidates and returns the latest date object + display text.
    """
    text_value = clean(s)
    if not text_value:
        return pd.NaT, ""

    candidates = []

    # Dates with day + month + year: 10 April 2028, 3 May 2028, 2 July 2028.
    for m in re.finditer(r"\b\d{1,2}\s+[A-Za-z]+\s+\d{4}\b", text_value, flags=re.I):
        candidates.append(m.group(0))

    # ISO-like dates.
    for m in re.finditer(r"\b\d{4}[-/]\d{1,2}[-/]\d{1,2}\b", text_value):
        candidates.append(m.group(0))

    # Month + year after Exp/Expired. Use day 1 for status calculation.
    for m in re.finditer(r"(?:exp|expired|berlaku|s\.?d\.?)\s*[:;()\-]*\s*([A-Za-z]+)\s+(\d{4})", text_value, flags=re.I):
        month_word, year = m.group(1), m.group(2)
        candidates.append(f"1 {month_word} {year}")

    parsed = []
    for candidate in candidates:
        dt = parse_date(candidate)
        if pd.notna(dt):
            parsed.append((pd.Timestamp(dt), show_date(candidate)))

    if not parsed:
        return pd.NaT, ""

    parsed.sort(key=lambda x: x[0])
    return parsed[-1]


def status_ska(dt):
    if pd.isna(dt): return "Tidak diketahui"
    return "Aktif" if dt >= pd.Timestamp.now().normalize() else "Expired"


def pdf_safe(s): return str(s).encode("latin-1", errors="ignore").decode("latin-1")


def build_pdf(df, filters, max_rows=120):
    pdf = FPDF(orientation="L", unit="mm", format="A4"); pdf.set_auto_page_break(True, 10); pdf.add_page()
    pdf.set_font("Helvetica", "B", 14); pdf.cell(0, 8, "Laporan Hasil Filter Personil", ln=True)
    pdf.set_font("Helvetica", "", 9); pdf.cell(0, 6, f"Dibuat: {datetime.now().strftime('%d-%m-%Y %H:%M')}", ln=True); pdf.cell(0, 6, f"Jumlah hasil: {len(df)} data", ln=True)
    active = {k:v for k,v in filters.items() if clean(v)}
    if active:
        pdf.ln(1); pdf.set_font("Helvetica", "B", 9); pdf.cell(0, 6, "Filter aktif:", ln=True); pdf.set_font("Helvetica", "", 8)
        for k,v in active.items(): pdf.cell(0, 5, pdf_safe(f"{k}: {v}"), ln=True)
    shown = df.head(max_rows).copy(); pdf.ln(3)
    if shown.empty: pdf.cell(0, 8, "Tidak ada data.", ln=True)
    else:
        cols = list(shown.columns); usable = 277; width = max(24, usable/max(1,len(cols)))
        if width * len(cols) > usable:
            cols = cols[:max(1, int(usable//24))]; shown = shown[cols]; width = usable/len(cols)
        pdf.set_font("Helvetica", "B", 7)
        for c in cols: pdf.cell(width, 7, pdf_safe(c[:22]), border=1)
        pdf.ln(7); pdf.set_font("Helvetica", "", 6.5)
        for _, row in shown.iterrows():
            for c in cols:
                v = pdf_safe(row.get(c, ""))
                if len(v) > 32: v = v[:29] + "..."
                pdf.cell(width, 6, v, border=1)
            pdf.ln(6)
    out = pdf.output(dest="S")
    return out.encode("latin-1", errors="ignore") if isinstance(out, str) else bytes(out)


def get_config():
    url, refresh = None, DEFAULT_REFRESH_SECONDS
    try:
        url = st.secrets.get("google_sheet", {}).get("xlsx_export_url")
        refresh = int(st.secrets.get("google_sheet", {}).get("refresh_seconds", DEFAULT_REFRESH_SECONDS))
    except Exception: pass
    return clean(url), max(30, refresh)

@st.cache_data(ttl=DEFAULT_REFRESH_SECONDS, show_spinner=False)
def download_xlsx(url):
    r = requests.get(url, timeout=90); r.raise_for_status()
    ctype = r.headers.get("content-type", "").lower()
    if "text/html" in ctype and b"<html" in r.content[:500].lower():
        raise RuntimeError("Google Sheet tidak bisa diakses sebagai XLSX. Pastikan sharing: Anyone with the link → Viewer, dan export URL benar.")
    return r.content


def read_workbook(b):
    return pd.read_excel(io.BytesIO(b), sheet_name=None, header=None, dtype=str, engine="openpyxl")


def read_hyperlinks(b):
    """Extract hyperlink targets from Google Sheets XLSX export.

    Returns:
        {sheet_name: {(data_row_index_after_header_reset, canonical_header): url}}
    """
    link_map = {}
    try:
        wb = load_workbook(io.BytesIO(b), read_only=False, data_only=True)
        for ws in wb.worksheets:
            sheet_links = {}
            rows = list(ws.iter_rows())
            header_idx = None
            headers = []
            for idx, row in enumerate(rows[:30]):
                headers = [norm_header(cell.value) for cell in row]
                if "NAMA" in headers:
                    header_idx = idx
                    break
            if header_idx is None:
                continue
            for ridx, row in enumerate(rows[header_idx + 1:], start=0):
                for cidx, cell in enumerate(row):
                    if not cell.hyperlink:
                        continue
                    target = clean(getattr(cell.hyperlink, "target", "") or getattr(cell.hyperlink, "location", ""))
                    if not target:
                        continue
                    canon = classify(headers[cidx] if cidx < len(headers) else "")
                    if canon:
                        sheet_links[(ridx, canon)] = target
            link_map[ws.title] = sheet_links
    except Exception:
        return {}
    return link_map


def find_header(df):
    for i in range(min(len(df), 30)):
        if "NAMA" in [norm_header(v) for v in df.iloc[i].tolist()]: return i
    return None


def skip_sheet(name):
    s = clean(name).lower()
    return any(k in s for k in BLACKLIST)


def process_workbook(raw_sheets: Dict[str, pd.DataFrame], link_map: Optional[Dict[str, Dict[Tuple[int, str], str]]] = None):
    rows, processed, skipped, notes_count = [], [], [], 0
    link_map = link_map or {}
    for sheet, raw in raw_sheets.items():
        if skip_sheet(sheet) or raw.empty: skipped.append(sheet); continue
        h = find_header(raw)
        if h is None: skipped.append(sheet); continue
        df = raw.iloc[h+1:].copy().reset_index(drop=True); df.columns = unique_cols(raw.iloc[h].tolist()); df = df.dropna(how="all")
        groups = {}
        for col in df.columns:
            canon = classify(col)
            if canon: groups.setdefault(canon, []).append(col)
        if "NAMA" not in groups: skipped.append(sheet); continue
        processed.append(sheet)
        sheet_links = link_map.get(sheet, {})
        for ridx, row in df.iterrows():
            nama = first(row, groups.get("NAMA", []))
            nama_link = clean(sheet_links.get((int(ridx), "NAMA"), ""))
            if not nama: continue
            vals = [clean(v) for v in row.tolist()]; notes = []
            kota = first(row, groups.get("KOTA/KABUPATEN", []))
            prov = first(row, groups.get("PROVINSI", []))
            dom_raw = first(row, groups.get("DOMISILI", []))
            # Kolom gabungan DOMISILI tidak lagi ditampilkan.
            # Jika Excel lama hanya punya kolom gabungan, nilai itu dipakai sebagai fallback untuk kota/provinsi.
            if dom_raw and not kota and not prov:
                parts = [clean(x) for x in re.split(r"\s*/\s*|\s*,\s*|\s+-\s+", dom_raw) if clean(x)]
                if len(parts) >= 2:
                    kota, prov = parts[0], parts[1]
                elif len(parts) == 1:
                    kota = parts[0]
            dom_search = " ".join([x for x in [kota, prov, dom_raw] if x])
            for col in groups.get("NO NIK", []):
                t = clean(row.get(col, ""))
                if t and not nik_of(t) and not any(ch.isdigit() for ch in t) and len(t) > 3 and not dom_raw:
                    kota = t; dom_search = " ".join([x for x in [kota, prov, dom_raw] if x]); notes.append("Kota/Kabupaten dipindahkan dari kolom NIK")
            nik = next((nik_of(row.get(c, "")) for c in groups.get("NO NIK", []) if nik_of(row.get(c, ""))), "")
            if not nik:
                for v in vals:
                    nik = nik_of(v)
                    if nik: notes.append("NIK ditemukan dari kolom lain"); break
            npwp = next((npwp_of(row.get(c, "")) for c in groups.get("NO NPWP", []) + groups.get("NPWP_RAW", []) if npwp_of(row.get(c, ""))), "")
            if not npwp:
                for v in vals:
                    npwp = npwp_of(v)
                    if npwp: notes.append("NPWP ditemukan dari kolom lain"); break
            email = next((email_of(row.get(c, "")) for c in groups.get("EMAIL", []) if email_of(row.get(c, ""))), "")
            if not email:
                for v in vals:
                    email = email_of(v)
                    if email: break
            phone = next((phone_of(row.get(c, "")) for c in groups.get("NO. TELP", []) if phone_of(row.get(c, ""))), "")
            if not phone:
                for col, v in zip(df.columns, vals):
                    if classify(col) in {"TAHUN LULUS IJAZAH","TGL EXPIRED SKA","BERLAKU SKA","JENIS IJAZAH","NO NIK","NO NPWP","NPWP_RAW"}: continue
                    phone = phone_of(v)
                    if phone: notes.append("Nomor telepon ditemukan dari kolom lain"); break
            sumber = next((source_clean(row.get(c, "")) for c in groups.get("SUMBER", []) if source_clean(row.get(c, ""))), "")
            if not sumber:
                for c in groups.get("EMAIL", []):
                    raw_email = clean(row.get(c, ""))
                    if raw_email and not email_of(raw_email) and not is_check(raw_email) and not phone_of(raw_email):
                        sumber = source_clean(raw_email)
                        if sumber: notes.append("Sumber data dipindahkan dari kolom email"); break
            keahlian_raw = first(row, groups.get("KEAHLIAN", []))
            exp_raw = first(row, groups.get("TGL EXPIRED SKA", [])); berlaku_raw = first(row, groups.get("BERLAKU SKA", [])); exp_obj = parse_date(exp_raw)
            if pd.isna(exp_obj) and berlaku_raw:
                exp_obj = parse_date(berlaku_raw)
                if pd.notna(exp_obj) and not exp_raw: exp_raw = berlaku_raw
            if pd.isna(exp_obj) and keahlian_raw:
                exp_obj, exp_from_skill = latest_expiry_from_text(keahlian_raw)
                if pd.notna(exp_obj):
                    exp_raw = exp_from_skill
                    notes.append("Tanggal expired SKA diambil dari kolom keahlian")
            jenis_raw = first_education(row, groups.get("JENIS IJAZAH", []))
            jenis_label = clean_education_label(jenis_raw) or jenis_raw
            tahun_raw = first(row, groups.get("TAHUN LULUS IJAZAH", []))
            tahun_from_education = False
            if not tahun_raw:
                tahun_raw = extract_years_from_education_text(jenis_raw)
                tahun_from_education = bool(tahun_raw)
                if tahun_from_education: notes.append("Tahun lulus diambil dari kolom ijazah/kelulusan")
            tahun_display = tahun_raw if tahun_from_education else show_date(tahun_raw)
            rows.append({"NO": first(row, groups.get("NO", [])), "NAMA": nama, "LINK PERSONIL": nama_link, "STRATA": first(row, groups.get("STRATA", [])), "JENIS IJAZAH": jenis_label, "TAHUN LULUS IJAZAH": tahun_display, "KEAHLIAN": keahlian_raw, "PENERBIT SKA/SKK": first(row, groups.get("PENERBIT SKA/SKK", [])), "BERLAKU SKA": show_date(berlaku_raw), "TGL EXPIRED SKA": show_date(exp_raw), "STATUS SKA": status_ska(exp_obj), "STATUS KERJA": first(row, groups.get("STATUS KERJA", [])), "KOTA/KABUPATEN": kota, "PROVINSI": prov, "NO NIK": nik, "NO NPWP": npwp, "NO. TELP": phone, "EMAIL": email, "SUMBER": sumber, "KATEGORI_ASAL": sheet, "CATATAN AUTO-CLEANING": "; ".join(dict.fromkeys(notes)), "_EXPIRED_DATE_OBJ": exp_obj})
            if notes: notes_count += 1
    if not rows:
        return pd.DataFrame(), {"processed_sheets": processed, "skipped_sheets": skipped, "notes_count": notes_count}
    out = pd.DataFrame(rows).fillna(""); out = out[out["NAMA"].astype(str).str.strip() != ""].reset_index(drop=True); out["NO"] = range(1, len(out)+1)
    out["_DOMISILI_SEARCH"] = (out["KOTA/KABUPATEN"].astype(str)+" "+out["PROVINSI"].astype(str)).map(norm_filter)
    # Strict pendidikan search.
    # JENIS IJAZAH is prioritized over STRATA for degree-level detection.
    edu_helpers = out.apply(
        lambda row: build_education_helpers(row.get("JENIS IJAZAH", ""), row.get("STRATA", "")),
        axis=1,
    )
    out["_PENDIDIKAN_SEARCH"] = edu_helpers.map(lambda x: x[0])
    out["_EDU_LEVELS"] = edu_helpers.map(lambda x: x[1])
    out["_KEAHLIAN_SEARCH"] = out["KEAHLIAN"].astype(str).map(norm_filter)
    out["_SEARCH_TEXT"] = out[MAIN_COLS].astype(str).agg(" ".join, axis=1).map(norm_filter)
    return out, {"processed_sheets": processed, "skipped_sheets": skipped, "notes_count": notes_count, "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

@st.cache_data(ttl=DEFAULT_REFRESH_SECONDS, show_spinner=False)
def load_live_data(url):
    content = download_xlsx(url)
    return process_workbook(read_workbook(content), read_hyperlinks(content))


def apply_filters(df, global_q, dom_q, skill_q, edu_q, year_q, publisher_q, status_q):
    if df.empty:
        return df
    r = df.copy()

    # Defensive helper columns. This prevents crashes if Streamlit cache still holds
    # a dataframe generated by an older app version.
    if "_DOMISILI_SEARCH" not in r.columns:
        r["_DOMISILI_SEARCH"] = (
            r.get("KOTA/KABUPATEN", pd.Series([""] * len(r), index=r.index)).astype(str)
            + " "
            + r.get("PROVINSI", pd.Series([""] * len(r), index=r.index)).astype(str)
        ).map(norm_filter)
    if "_PENDIDIKAN_SEARCH" not in r.columns or "_EDU_LEVELS" not in r.columns:
        edu_helpers = r.apply(
            lambda row: build_education_helpers(row.get("JENIS IJAZAH", ""), row.get("STRATA", "")),
            axis=1,
        )
        r["_PENDIDIKAN_SEARCH"] = edu_helpers.map(lambda x: x[0])
        r["_EDU_LEVELS"] = edu_helpers.map(lambda x: x[1])
    if "_KEAHLIAN_SEARCH" not in r.columns:
        r["_KEAHLIAN_SEARCH"] = r.get("KEAHLIAN", pd.Series([""] * len(r), index=r.index)).astype(str).map(norm_filter)
    if "_SEARCH_TEXT" not in r.columns:
        safe_cols = [c for c in MAIN_COLS if c in r.columns]
        r["_SEARCH_TEXT"] = r[safe_cols].astype(str).agg(" ".join, axis=1).map(norm_filter) if safe_cols else ""

    r = r[contains_global(r, global_q)]
    if clean(dom_q):
        r = r[contains_groups(r["_DOMISILI_SEARCH"], dom_q)]
    terms = [simplify_skill(t) or t for t in split_terms(skill_q)]
    if terms:
        r = r[contains_any(r["_KEAHLIAN_SEARCH"], terms)]
    if clean(edu_q):
        # Smart education filter:
        # "Sipil D4" matches "D4 Sipil".
        # "Sipil D4 D3" becomes (Sipil+D4) OR (Sipil+D3).
        # "S1, D3" or "S1 serta D3" becomes S1 OR D3.
        r = r[contains_education_query(r, edu_q)]
    if clean(year_q):
        r = r[contains_groups(r["TAHUN LULUS IJAZAH"].astype(str).map(norm_filter), year_q)]
    if clean(publisher_q):
        r = r[contains_groups(r["PENERBIT SKA/SKK"], publisher_q)]
    if status_q and status_q != "Semua": r = r[r["STATUS SKA"].astype(str).str.lower() == status_q.lower()]
    return r.reset_index(drop=True)


def prepare_display(df, cols):
    visible = [c for c in cols if c in df.columns and not c.startswith("_")] or [c for c in DEFAULT_COLS if c in df.columns]
    return df[visible].copy().fillna("")


def render_header(n, update, refresh):
    st.markdown(
        '<div class="main-header">'
        '<div class="header-title">Portal Filter Tenaga Ahli</div>'
        '<div class="header-subtitle">Pencarian dan sortir personil berdasarkan kota/kabupaten, provinsi, keahlian/SKK, status SKA, dan pendidikan.</div>'
        '</div>',
        unsafe_allow_html=True,
    )
    html = (
        f'<div class="metric-grid">'
        f'<div class="metric-card"><div class="metric-label">Total Data</div><div class="metric-value">{n:,}</div><div class="metric-note">record terbaca dari Google Sheet</div></div>'
        f'<div class="metric-card"><div class="metric-label">Data Terakhir Dibaca</div><div class="metric-value" style="font-size:18px; margin-top:14px;">{update or "-"}</div><div class="metric-note">refresh otomatis tiap {refresh} detik</div></div>'
        f'</div>'
    )
    st.markdown(html.replace(",", "."), unsafe_allow_html=True)


def main():
    url, refresh = get_config()
    if not url:
        st.markdown('<div class="main-header"><div class="header-title">Portal Filter Tenaga Ahli</div><div class="header-subtitle">Google Sheet belum dikonfigurasi</div></div>', unsafe_allow_html=True)
        st.markdown('<div class="notice-box">Tambahkan ini di Streamlit Cloud → Settings → Secrets:<br><br>[google_sheet]<br>xlsx_export_url = "https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/export?format=xlsx"<br>refresh_seconds = 60</div>', unsafe_allow_html=True)
        st.stop()
    try:
        with st.spinner("Mengambil dan membersihkan data dari Google Sheets..."):
            df, meta = load_live_data(url)
    except Exception as e:
        st.error(str(e)); st.stop()
    render_header(len(df), meta.get("generated_at", ""), refresh)
    if df.empty: st.warning("Data kosong. Pastikan Google Sheet bisa diakses dan masih berisi kolom NAMA."); st.stop()
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Filter Data</div><div class="section-caption">Masukkan kata kunci lalu klik Cari. Contoh: Manado, Sulut, Bali atau Sipil D4 D3.</div>', unsafe_allow_html=True)
    q_col, cari_col, refresh_col = st.columns([4.2, 1, 1])
    with q_col:
        global_q = st.text_input("Pencarian umum", placeholder="Cari nama, NIK, NPWP, kota/provinsi, keahlian...", label_visibility="collapsed", key="global_q")
    with cari_col:
        search_clicked = st.button("Cari", type="primary", use_container_width=True)
    with refresh_col:
        if st.button("Ambil Data", use_container_width=True):
            st.cache_data.clear()
            st.rerun()
    c1, c2, c3, c4 = st.columns([1.25, 1.25, 1.25, 1])
    with c1: dom_q = st.text_input("Kota/Provinsi", placeholder="Manado, Sulut, Bali", key="dom_q")
    with c2: skill_q = st.text_input("Keahlian / SKK", placeholder="Jalan, Gedung, Arsitek", key="skill_q")
    with c3: edu_q = st.text_input("Pendidikan / Ijazah", placeholder="Sipil D4 D3 / Sipil D IV / S1, D3", key="edu_q")
    with c4: status_q = st.selectbox("Status SKA", ["Semua", "Aktif", "Expired", "Tidak diketahui"], key="status_q")
    c5, c6 = st.columns([1, 1])
    with c5: year_q = st.text_input("Tahun Lulus", placeholder="2017, 2018", key="year_q")
    with c6: publisher_q = st.text_input("Penerbit SKA/SKK", placeholder="LPJK, PUPR", key="publisher_q")
    st.markdown('</div>', unsafe_allow_html=True)
    filtered = apply_filters(df, global_q, dom_q, skill_q, edu_q, year_q, publisher_q, status_q)
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown(f'<div class="section-title">Hasil Filter</div><div class="section-caption">{len(filtered):,} data dari {len(df):,} record</div>'.replace(',', '.'), unsafe_allow_html=True)
    available = [c for c in MAIN_COLS + ["CATATAN AUTO-CLEANING"] if c in filtered.columns]
    selected = st.multiselect("Kolom tampilan", options=available, default=[c for c in DEFAULT_COLS if c in available])
    display_df = prepare_display(filtered, selected)
    column_config = {}
    if "LINK PERSONIL" in display_df.columns:
        column_config["LINK PERSONIL"] = st.column_config.LinkColumn("Link Personil", display_text="Buka")
    st.dataframe(display_df.head(500), use_container_width=True, height=520, column_config=column_config)
    if len(display_df) > 500: st.caption("Tabel menampilkan 500 baris pertama. Gunakan Download CSV untuk seluruh data.")
    filters = {"Pencarian umum":global_q,"Kota/Provinsi":dom_q,"Keahlian/SKK":skill_q,"Pendidikan/Ijazah":edu_q,"Tahun Lulus":year_q,"Penerbit":publisher_q,"Status SKA": status_q if status_q != "Semua" else ""}
    d1, d2 = st.columns(2)
    with d1: st.download_button("Download CSV", data=display_df.to_csv(index=False).encode("utf-8-sig"), file_name=f"hasil_filter_personil_{datetime.now().strftime('%Y%m%d_%H%M')}.csv", mime="text/csv", use_container_width=True)
    with d2: st.download_button("Download PDF", data=build_pdf(display_df, filters), file_name=f"hasil_filter_personil_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf", mime="application/pdf", use_container_width=True)
    with st.expander("Info sumber data"):
        st.write({"sheet_diproses": meta.get("processed_sheets", []), "sheet_dilewati": meta.get("skipped_sheets", []), "baris_auto_fix": meta.get("notes_count", 0), "update_terakhir": meta.get("generated_at", "")})
    st.markdown('</div>', unsafe_allow_html=True)

if __name__ == "__main__": main()
